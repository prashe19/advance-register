import os
import re
import secrets
import sqlite3
import smtplib
from email.mime.text import MIMEText
from datetime import date, datetime, timedelta
from functools import wraps
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from flask import Flask, request, jsonify, g, send_file
from flask_cors import CORS
from dotenv import load_dotenv
from werkzeug.security import generate_password_hash, check_password_hash

load_dotenv()

DB_PATH = os.environ.get(
    "DATABASE_PATH",
    os.path.join(os.path.dirname(__file__), "advance_register.db"),
)

app = Flask(__name__)
# CORS_ORIGINS: comma-separated list of allowed frontend origins in production
# (e.g. "https://advances.yourinstitute.edu"). Defaults to "*" only because
# that's convenient for local development -- set this explicitly before
# deploying anywhere reachable from outside your own machine.
_cors_origins = os.environ.get("CORS_ORIGINS", "*")
_cors_origins = [o.strip() for o in _cors_origins.split(",")] if _cors_origins != "*" else "*"
CORS(app, resources={r"/api/*": {"origins": _cors_origins}}, allow_headers="*", expose_headers="*")

# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def _column_exists(conn, table, column):
    cols = [r[1] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()]
    return column in cols


# Default set of department codes/names offered in the "Department" dropdown
# and used to validate advances on create. Edit this list (or the
# `departments` table directly) to match your institute's departments --
# it's only used to seed an empty table, so changes here won't touch rows
# that already exist.
DEFAULT_DEPARTMENTS = [
    ("AE", "Aerospace"),
    ("AM", "Applied Mechanics"),
    ("BT", "Bio Technology"),
    ("CE", "Civil"),
    ("CH", "Chemical"),
    ("CS", "Computer Science"),
    ("CY", "Chemistry"),
    ("DA", "Data Science"),
    ("ED", "Engineering Design"),
    ("EE", "Electrical"),
    ("ID", "Inter Disciplinary"),
    ("MA", "Mathematics"),
    ("ME", "Mechanical"),
    ("MM", "Metallurgy"),
    ("MS", "Management Studies"),
    ("OE", "Ocean"),
    ("PH", "Physics"),
    ("HS", "Humanities"),
]

# Number of days after a conference ends that an advance is expected to be
# settled by. Used to compute `due_date` / `is_overdue` on pending advances.
SETTLEMENT_GRACE_DAYS = 15

ADVANCE_STATUSES = (
    "pending",
    "closed",
    "closed_payment_due",
    "closed_refund_due",
    "payment_made",
    "refund_received",
)

# Convenience groupings for the Advances-tab status filter and reports --
# "needs_action" is anything settled-but-not-yet-paid-out; "settled" is
# anything fully done, however it got there.
STATUS_GROUPS = {
    "needs_action": ("closed_payment_due", "closed_refund_due"),
    "settled": ("closed", "payment_made", "refund_received"),
}

# Fields on `advances` a caller may set directly (used by update_advance and
# by the history diff helper). Anything else is derived/internal.
ADVANCE_HISTORY_FIELDS = (
    "name", "identifier", "email", "department", "amount", "date_given",
    "conf_start_date", "conf_end_date", "purpose", "status", "date_settled",
    "amount_spent", "amount_refundable", "amount_paid", "date_paid", "payment_ref",
    "amount_received", "date_received", "receipt_ref", "voucher_no",
)


def _migrate_advances_table(conn):
    """The original `advances` table only supported a two-state
    pending/adjusted workflow (enforced via a CHECK constraint) and had no
    department / voucher / settlement columns. SQLite can't ALTER a CHECK
    constraint or add columns with rich defaults after the fact, so if we
    detect the old shape we rebuild the table under a new name, copy the
    existing rows across, and swap it in. Safe to run every startup -- it's
    a no-op once the table is already in the new shape.
    """
    if not _column_exists(conn, "advances", "department"):
        conn.execute("ALTER TABLE advances RENAME TO advances_legacy")
        conn.execute(
            f"""
            CREATE TABLE advances (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category TEXT NOT NULL CHECK(category IN ('student','faculty')),
                identifier TEXT,
                name TEXT NOT NULL,
                email TEXT,
                department TEXT,
                amount REAL NOT NULL,
                date_given TEXT NOT NULL,
                conf_start_date TEXT,
                conf_end_date TEXT,
                purpose TEXT,
                status TEXT NOT NULL DEFAULT 'pending'
                    CHECK(status IN {ADVANCE_STATUSES!r}),
                date_adjusted TEXT,
                amount_spent REAL,
                amount_refundable REAL,
                amount_paid REAL,
                date_paid TEXT,
                payment_ref TEXT,
                amount_received REAL,
                date_received TEXT,
                receipt_ref TEXT,
                voucher_no TEXT,
                financial_year TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        old_cols = [r[1] for r in conn.execute("PRAGMA table_info(advances_legacy)").fetchall()]
        common = [c for c in old_cols if c not in ("id", "status")]
        col_list = ", ".join(common)
        # Older DBs may still have rows in the two-state 'adjusted' status,
        # which no longer exists -- fold those into 'closed' on the way in,
        # since no spend figure was ever recorded for them to reconcile.
        conn.execute(
            f"INSERT INTO advances (id, {col_list}, status) "
            f"SELECT id, {col_list}, CASE WHEN status='adjusted' THEN 'closed' ELSE status END "
            f"FROM advances_legacy"
        )
        conn.execute("DROP TABLE advances_legacy")

        # Backfill voucher_no / financial_year for rows that predate them,
        # numbering each financial year's vouchers 1, 2, 3... in id order.
        rows = conn.execute("SELECT id, date_given FROM advances WHERE voucher_no IS NULL ORDER BY id").fetchall()
        fy_counters = {}
        for adv_id, date_given in rows:
            fy = compute_financial_year(date_given)
            fy_counters[fy] = fy_counters.get(fy, 0) + 1
            voucher_no = f"ADV/{fy}/{fy_counters[fy]:04d}"
            conn.execute("UPDATE advances SET voucher_no = ?, financial_year = ? WHERE id = ?", (voucher_no, fy, adv_id))


def _migrate_to_students_only(conn):
    """The register used to cover both students and faculty (via a `category`
    column) and had a manual 'adjusted' status that closed a pending advance
    without recording what was actually spent. The register is students-only
    now, and every close goes through the settle flow instead, which records
    amount_spent and lands on 'closed' (spent == given), 'closed_payment_due'
    or 'closed_refund_due'. SQLite can't drop a column that's covered by a
    CHECK constraint in place, so if the old shape is detected we rebuild the
    table, translating any legacy 'adjusted' rows to 'closed' (nothing was
    ever recorded for them to reconcile) and dropping `category` entirely.
    Safe to run every startup -- a no-op once already migrated.
    """
    if _column_exists(conn, "advances", "category"):
        conn.execute("ALTER TABLE advances RENAME TO advances_legacy")
        conn.execute(
            f"""
            CREATE TABLE advances (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                identifier TEXT,
                name TEXT NOT NULL,
                email TEXT,
                department TEXT,
                amount REAL NOT NULL,
                date_given TEXT NOT NULL,
                conf_start_date TEXT,
                conf_end_date TEXT,
                purpose TEXT,
                status TEXT NOT NULL DEFAULT 'pending' CHECK(status IN {ADVANCE_STATUSES!r}),
                date_settled TEXT,
                amount_spent REAL,
                amount_refundable REAL,
                amount_paid REAL,
                date_paid TEXT,
                payment_ref TEXT,
                amount_received REAL,
                date_received TEXT,
                receipt_ref TEXT,
                voucher_no TEXT,
                financial_year TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        old_cols = [r[1] for r in conn.execute("PRAGMA table_info(advances_legacy)").fetchall()]
        # category is dropped entirely; status and date_adjusted are translated
        # explicitly below rather than copied verbatim.
        common = [c for c in old_cols if c not in ("id", "category", "status", "date_adjusted")]
        col_list = ", ".join(common)
        conn.execute(
            f"INSERT INTO advances (id, {col_list}, status, date_settled) "
            f"SELECT id, {col_list}, "
            f"CASE WHEN status='adjusted' THEN 'closed' ELSE status END, "
            f"date_adjusted "
            f"FROM advances_legacy"
        )
        conn.execute("DROP TABLE advances_legacy")
    elif _column_exists(conn, "advances", "date_adjusted") and not _column_exists(conn, "advances", "date_settled"):
        # Shouldn't normally happen (category and date_adjusted were removed
        # together), but handle it defensively.
        conn.execute("ALTER TABLE advances RENAME COLUMN date_adjusted TO date_settled")


def init_db():
    conn = sqlite3.connect(DB_PATH)

    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS advances (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            identifier TEXT,
            name TEXT NOT NULL,
            email TEXT,
            department TEXT,
            amount REAL NOT NULL,
            date_given TEXT NOT NULL,
            conf_start_date TEXT,
            conf_end_date TEXT,
            purpose TEXT,
            status TEXT NOT NULL DEFAULT 'pending' CHECK(status IN {ADVANCE_STATUSES!r}),
            date_settled TEXT,
            amount_spent REAL,
            amount_refundable REAL,
            amount_paid REAL,
            date_paid TEXT,
            payment_ref TEXT,
            amount_received REAL,
            date_received TEXT,
            receipt_ref TEXT,
            voucher_no TEXT,
            financial_year TEXT,
            created_at TEXT NOT NULL
        )
        """
    )
    _migrate_advances_table(conn)
    _migrate_to_students_only(conn)
    if not _column_exists(conn, "advances", "identifier"):
        conn.execute("ALTER TABLE advances ADD COLUMN identifier TEXT")

    # NOTE: uniqueness of roll no is enforced in the application layer (see
    # identifier_conflict()), not as a hard DB constraint -- a student can and
    # should have multiple advance rows over time. What must stay unique is
    # which *name* an identifier belongs to, not the number of rows.
    conn.execute("DROP INDEX IF EXISTS idx_advances_category_identifier")

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS departments (
            code TEXT PRIMARY KEY,
            name TEXT NOT NULL
        )
        """
    )
    if conn.execute("SELECT COUNT(*) FROM departments").fetchone()[0] == 0:
        conn.executemany("INSERT INTO departments (code, name) VALUES (?, ?)", DEFAULT_DEPARTMENTS)
    else:
        # One-time cleanup: earlier test runs may have seeded a generic
        # placeholder list (CSE/ECE/...) before the real department codes
        # were known. If the table still looks like that placeholder set
        # (none of the real codes present, and it's not referenced by any
        # advance yet), swap it out. Leaves any table an admin has already
        # customized or that's in real use untouched.
        existing_codes = {r[0] for r in conn.execute("SELECT code FROM departments").fetchall()}
        real_codes = {code for code, _ in DEFAULT_DEPARTMENTS}
        in_use = conn.execute("SELECT COUNT(*) FROM advances WHERE department IS NOT NULL").fetchone()[0]
        if not (existing_codes & real_codes) and in_use == 0:
            conn.execute("DELETE FROM departments")
            conn.executemany("INSERT INTO departments (code, name) VALUES (?, ?)", DEFAULT_DEPARTMENTS)

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS advance_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            advance_id INTEGER NOT NULL,
            field TEXT NOT NULL,
            old_value TEXT,
            new_value TEXT,
            changed_by TEXT,
            changed_at TEXT NOT NULL
        )
        """
    )

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL CHECK(role IN ('admin','staff')),
            email TEXT,
            created_at TEXT NOT NULL
        )
        """
    )
    if not _column_exists(conn, "users", "email"):
        conn.execute("ALTER TABLE users ADD COLUMN email TEXT")
    if not _column_exists(conn, "advances", "last_reminded_at"):
        conn.execute("ALTER TABLE advances ADD COLUMN last_reminded_at TEXT")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS sessions (
            token TEXT PRIMARY KEY,
            user_id INTEGER NOT NULL,
            created_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS password_resets (
            code TEXT PRIMARY KEY,
            user_id INTEGER NOT NULL,
            expires_at TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
        """
    )
    conn.commit()

    existing = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if existing == 0:
        bootstrap_user = os.environ.get("ADMIN_BOOTSTRAP_USERNAME", "admin")
        bootstrap_pass = os.environ.get("ADMIN_BOOTSTRAP_PASSWORD", "ChangeMe123!")
        conn.execute(
            "INSERT INTO users (username, password_hash, role, created_at) VALUES (?, ?, 'admin', ?)",
            (bootstrap_user, generate_password_hash(bootstrap_pass), datetime.utcnow().isoformat()),
        )
        conn.commit()
        print(
            f"\n[advance-register] Created first admin account -> username: '{bootstrap_user}'  "
            f"password: '{bootstrap_pass}'\n"
            "[advance-register] Log in and create real accounts, or set ADMIN_BOOTSTRAP_USERNAME / "
            "ADMIN_BOOTSTRAP_PASSWORD in backend/.env before first run.\n"
        )

    conn.close()


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

def get_current_user():
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        return None
    token = auth[7:].strip()
    if not token:
        return None
    db = get_db()
    row = db.execute(
        "SELECT users.id AS id, users.username AS username, users.role AS role "
        "FROM sessions JOIN users ON sessions.user_id = users.id "
        "WHERE sessions.token = ?",
        (token,),
    ).fetchone()
    return row


def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        user = get_current_user()
        if not user:
            return jsonify({"error": "Please log in to continue."}), 401
        g.current_user = user
        return f(*args, **kwargs)
    return wrapper


def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        user = get_current_user()
        if not user:
            return jsonify({"error": "Please log in to continue."}), 401
        if user["role"] != "admin":
            return jsonify({"error": "Admin access required."}), 403
        g.current_user = user
        return f(*args, **kwargs)
    return wrapper


# ---------------------------------------------------------------------------
# Login rate limiting
#
# A simple in-memory sliding-window limiter: N failed attempts per key within
# a window locks that key out for a cooldown period. Keyed on IP+username
# together, so it can't be used to lock a real user out by spamming their
# username from many IPs, nor does one noisy IP block every account.
#
# This is intentionally in-process (no Redis/external store) to keep the
# deployment simple -- fine for a single-process gunicorn worker. If you
# scale to multiple worker processes behind nginx, move this to a shared
# store (Redis) or front the app with rate limiting at the nginx layer
# instead, since each worker would otherwise track attempts separately.
# ---------------------------------------------------------------------------

LOGIN_MAX_ATTEMPTS = 5
LOGIN_WINDOW_SECONDS = 15 * 60
LOGIN_LOCKOUT_SECONDS = 15 * 60

_login_attempts = {}  # key -> list[timestamp]
_login_lockouts = {}  # key -> lockout_until_timestamp


def _login_rate_key():
    ip = request.headers.get("X-Forwarded-For", request.remote_addr or "unknown").split(",")[0].strip()
    username = (request.get_json(silent=True) or {}).get("username", "").strip().lower()
    return f"{ip}:{username}"


def _login_rate_check():
    """Returns None if allowed, or an (response, status) tuple if blocked."""
    key = _login_rate_key()
    now = datetime.utcnow().timestamp()

    lockout_until = _login_lockouts.get(key)
    if lockout_until and now < lockout_until:
        wait_minutes = max(1, int((lockout_until - now) // 60) + 1)
        return jsonify({
            "error": f"Too many failed attempts. Try again in about {wait_minutes} minute(s)."
        }), 429

    return None


def _login_rate_record_failure():
    key = _login_rate_key()
    now = datetime.utcnow().timestamp()
    attempts = [t for t in _login_attempts.get(key, []) if now - t < LOGIN_WINDOW_SECONDS]
    attempts.append(now)
    _login_attempts[key] = attempts
    if len(attempts) >= LOGIN_MAX_ATTEMPTS:
        _login_lockouts[key] = now + LOGIN_LOCKOUT_SECONDS
        _login_attempts[key] = []


def _login_rate_clear():
    key = _login_rate_key()
    _login_attempts.pop(key, None)
    _login_lockouts.pop(key, None)


@app.route("/api/login", methods=["POST"])
def login():
    blocked = _login_rate_check()
    if blocked:
        return blocked

    data = request.get_json(force=True)
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    if not username or not password:
        return jsonify({"error": "Username and password are required."}), 400

    db = get_db()
    row = db.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
    if not row or not check_password_hash(row["password_hash"], password):
        _login_rate_record_failure()
        return jsonify({"error": "Incorrect username or password."}), 401

    _login_rate_clear()
    token = secrets.token_hex(32)
    db.execute(
        "INSERT INTO sessions (token, user_id, created_at) VALUES (?, ?, ?)",
        (token, row["id"], datetime.utcnow().isoformat()),
    )
    db.commit()
    return jsonify({"token": token, "username": row["username"], "role": row["role"]})


@app.route("/api/logout", methods=["POST"])
@login_required
def logout():
    auth = request.headers.get("Authorization", "")
    token = auth[7:].strip()
    db = get_db()
    db.execute("DELETE FROM sessions WHERE token = ?", (token,))
    db.commit()
    return jsonify({"loggedOut": True})


@app.route("/api/me", methods=["GET"])
@login_required
def me():
    db = get_db()
    row = db.execute("SELECT username, role, email FROM users WHERE id = ?", (g.current_user["id"],)).fetchone()
    return jsonify(dict(row))


@app.route("/api/users", methods=["GET"])
@admin_required
def list_users():
    db = get_db()
    rows = db.execute("SELECT id, username, role, email, created_at FROM users ORDER BY created_at").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/users", methods=["POST"])
@admin_required
def create_user():
    data = request.get_json(force=True)
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    role = data.get("role")
    email = (data.get("email") or "").strip()

    if role not in ("admin", "staff"):
        return jsonify({"error": "role must be 'admin' or 'staff'"}), 400
    if not username or not password:
        return jsonify({"error": "username and password are required"}), 400
    if len(password) < 6:
        return jsonify({"error": "password must be at least 6 characters"}), 400

    db = get_db()
    existing = db.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
    if existing:
        return jsonify({"error": "That username is already taken."}), 409

    cur = db.execute(
        "INSERT INTO users (username, password_hash, role, email, created_at) VALUES (?, ?, ?, ?, ?)",
        (username, generate_password_hash(password), role, email, datetime.utcnow().isoformat()),
    )
    db.commit()
    return jsonify({"id": cur.lastrowid, "username": username, "role": role, "email": email}), 201


@app.route("/api/users/<int:user_id>", methods=["PUT"])
@admin_required
def update_user(user_id):
    data = request.get_json(force=True)
    db = get_db()
    row = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404

    fields = {}
    if "email" in data:
        fields["email"] = (data.get("email") or "").strip()
    if "role" in data:
        if data["role"] not in ("admin", "staff"):
            return jsonify({"error": "role must be 'admin' or 'staff'"}), 400
        fields["role"] = data["role"]

    if not fields:
        return jsonify({"error": "no editable fields provided"}), 400

    set_clause = ", ".join(f"{k} = ?" for k in fields)
    params = list(fields.values()) + [user_id]
    db.execute(f"UPDATE users SET {set_clause} WHERE id = ?", params)
    db.commit()
    updated = db.execute("SELECT id, username, role, email FROM users WHERE id = ?", (user_id,)).fetchone()
    return jsonify(dict(updated))


@app.route("/api/users/<int:user_id>", methods=["DELETE"])
@admin_required
def delete_user(user_id):
    if user_id == g.current_user["id"]:
        return jsonify({"error": "You can't delete the account you're logged in as."}), 400
    db = get_db()
    db.execute("DELETE FROM sessions WHERE user_id = ?", (user_id,))
    db.execute("DELETE FROM password_resets WHERE user_id = ?", (user_id,))
    db.execute("DELETE FROM users WHERE id = ?", (user_id,))
    db.commit()
    return jsonify({"deleted": True})


@app.route("/api/users/<int:user_id>/reset-password", methods=["POST"])
@admin_required
def admin_reset_password(user_id):
    """Admin sets a new password for someone directly -- works even without email/SMTP configured."""
    data = request.get_json(force=True)
    new_password = data.get("new_password") or ""
    if len(new_password) < 6:
        return jsonify({"error": "password must be at least 6 characters"}), 400

    db = get_db()
    row = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404

    db.execute("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(new_password), user_id))
    db.execute("DELETE FROM sessions WHERE user_id = ?", (user_id,))  # force re-login everywhere
    db.commit()
    return jsonify({"reset": True, "username": row["username"]})


@app.route("/api/me/password", methods=["POST"])
@login_required
def change_own_password():
    data = request.get_json(force=True)
    current_password = data.get("current_password") or ""
    new_password = data.get("new_password") or ""
    if len(new_password) < 6:
        return jsonify({"error": "new password must be at least 6 characters"}), 400

    db = get_db()
    row = db.execute("SELECT * FROM users WHERE id = ?", (g.current_user["id"],)).fetchone()
    if not check_password_hash(row["password_hash"], current_password):
        return jsonify({"error": "Current password is incorrect."}), 401

    db.execute("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(new_password), row["id"]))
    db.commit()
    return jsonify({"changed": True})


def build_reset_email(code):
    subject = "Advance Register — password reset code"
    body = (
        f"Your password reset code is: {code}\n\n"
        "This code expires in 30 minutes. Enter it on the 'Forgot password' screen in the "
        "Advance Register app along with your new password.\n\n"
        "If you didn't request this, you can ignore this email."
    )
    return subject, body


@app.route("/api/forgot-password", methods=["POST"])
def forgot_password():
    data = request.get_json(force=True)
    username = (data.get("username") or "").strip()
    if not username:
        return jsonify({"error": "username is required"}), 400

    db = get_db()
    row = db.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()

    generic_ok = {"message": "If that account exists and has an email on file, a reset code has been sent."}

    if not row:
        return jsonify(generic_ok)
    if not row["email"]:
        return jsonify({"error": f"The account '{username}' has no email on file. Ask an admin to reset your password instead."}), 400
    if not smtp_configured():
        return jsonify({"error": "Email isn't configured on this server yet. Ask an admin to reset your password instead."}), 500

    code = "".join(secrets.choice("ABCDEFGHJKLMNPQRSTUVWXYZ23456789") for _ in range(8))
    expires_at = (datetime.utcnow().timestamp() + 30 * 60)
    db.execute(
        "INSERT INTO password_resets (code, user_id, expires_at, created_at) VALUES (?, ?, ?, ?)",
        (code, row["id"], datetime.utcfromtimestamp(expires_at).isoformat(), datetime.utcnow().isoformat()),
    )
    db.commit()

    subject, body = build_reset_email(code)
    ok, err = send_email(row["email"], subject, body)
    if not ok:
        return jsonify({"error": err}), 502

    return jsonify(generic_ok)


@app.route("/api/reset-password", methods=["POST"])
def reset_password_with_code():
    data = request.get_json(force=True)
    username = (data.get("username") or "").strip()
    code = (data.get("code") or "").strip().upper()
    new_password = data.get("new_password") or ""

    if not username or not code or not new_password:
        return jsonify({"error": "username, code and new_password are required"}), 400
    if len(new_password) < 6:
        return jsonify({"error": "new password must be at least 6 characters"}), 400

    db = get_db()
    user_row = db.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
    if not user_row:
        return jsonify({"error": "Invalid or expired code."}), 400

    reset_row = db.execute(
        "SELECT * FROM password_resets WHERE code = ? AND user_id = ?", (code, user_row["id"])
    ).fetchone()
    if not reset_row:
        return jsonify({"error": "Invalid or expired code."}), 400
    if datetime.fromisoformat(reset_row["expires_at"]) < datetime.utcnow():
        db.execute("DELETE FROM password_resets WHERE code = ?", (code,))
        db.commit()
        return jsonify({"error": "This code has expired. Request a new one."}), 400

    db.execute("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(new_password), user_row["id"]))
    db.execute("DELETE FROM password_resets WHERE user_id = ?", (user_row["id"],))
    db.execute("DELETE FROM sessions WHERE user_id = ?", (user_row["id"],))
    db.commit()
    return jsonify({"reset": True})


# ---------------------------------------------------------------------------
# Business rule: a student may not receive a new advance while one of their
# existing advances is still pending. Matched by roll number (identifier)
# rather than name.
# ---------------------------------------------------------------------------

def student_has_pending(db, identifier):
    return db.execute(
        "SELECT * FROM advances WHERE status='pending' "
        "AND lower(trim(identifier)) = lower(trim(?))",
        (identifier,),
    ).fetchone()


def identifier_conflict(db, identifier, name, exclude_id=None):
    sql = (
        "SELECT * FROM advances WHERE lower(trim(identifier)) = lower(trim(?)) "
        "AND lower(trim(name)) != lower(trim(?))"
    )
    params = [identifier, name]
    if exclude_id is not None:
        sql += " AND id != ?"
        params.append(exclude_id)
    sql += " LIMIT 1"
    return db.execute(sql, params).fetchone()


def compute_financial_year(date_given):
    """Indian financial year: 1 Apr -> 31 Mar, formatted like '2025-26'."""
    d = date.fromisoformat(date_given) if isinstance(date_given, str) else date_given
    start_year = d.year if d.month >= 4 else d.year - 1
    return f"{start_year}-{str(start_year + 1)[2:]}"


def generate_voucher_no(db, financial_year):
    seq = db.execute(
        "SELECT COUNT(*) FROM advances WHERE financial_year = ?", (financial_year,)
    ).fetchone()[0] + 1
    return f"ADV/{financial_year}/{seq:04d}"


def voucher_conflict(db, voucher_no, exclude_id=None):
    sql = "SELECT * FROM advances WHERE lower(trim(voucher_no)) = lower(trim(?))"
    params = [voucher_no]
    if exclude_id is not None:
        sql += " AND id != ?"
        params.append(exclude_id)
    sql += " LIMIT 1"
    return db.execute(sql, params).fetchone()


@app.route("/api/next-voucher", methods=["GET"])
@login_required
def next_voucher_preview():
    """
    What voucher number *would* be auto-assigned right now for this
    date_given, without reserving or writing anything -- generate_voucher_no
    is a plain count of existing rows for that financial year, so calling it
    here for a live preview is side-effect-free. Used to grey-fill the
    voucher field on the New Advance form before the person decides whether
    to override it.
    """
    date_given = request.args.get("date_given") or date.today().isoformat()
    try:
        financial_year = compute_financial_year(date_given)
    except ValueError:
        return jsonify({"error": "date_given must be a valid date (YYYY-MM-DD)"}), 400
    db = get_db()
    return jsonify({"voucher_no": generate_voucher_no(db, financial_year)})


def record_history(db, advance_id, changed_by, changes):
    """changes: iterable of (field, old_value, new_value). Skips no-op changes."""
    now = datetime.utcnow().isoformat()
    for field, old_value, new_value in changes:
        old_str = None if old_value is None else str(old_value)
        new_str = None if new_value is None else str(new_value)
        if old_str == new_str:
            continue
        db.execute(
            "INSERT INTO advance_history (advance_id, field, old_value, new_value, changed_by, changed_at) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (advance_id, field, old_str, new_str, changed_by, now),
        )


def _settlement_info(d):
    """Returns (due_date_iso_or_None, is_overdue, days_overdue) for a row dict.
    Only 'pending' advances can be overdue -- once settled the clock stops,
    regardless of how long settlement itself took.
    """
    conf_end = d.get("conf_end_date")
    if not conf_end or d.get("status") != "pending":
        return None, False, 0
    due = date.fromisoformat(conf_end) + timedelta(days=SETTLEMENT_GRACE_DAYS)
    overdue_days = (date.today() - due).days
    if overdue_days > 0:
        return due.isoformat(), True, overdue_days
    return due.isoformat(), False, 0


def row_to_dict(row):
    d = dict(row)
    if d.get("status") == "pending":
        d["days_outstanding"] = (date.today() - date.fromisoformat(d["date_given"])).days
    elif d.get("date_settled"):
        d["days_outstanding"] = (
            date.fromisoformat(d["date_settled"]) - date.fromisoformat(d["date_given"])
        ).days

    due_date, is_overdue, days_overdue = _settlement_info(d)
    d["due_date"] = due_date
    d["is_overdue"] = is_overdue
    d["days_overdue"] = days_overdue
    return d


# ---------------------------------------------------------------------------
# Advance routes (all require login)
# ---------------------------------------------------------------------------

SORT_COLUMNS = {
    "date": "date_given",
    "amount": "amount",
    "department": "department",
}


@app.route("/api/departments", methods=["GET"])
@login_required
def list_departments():
    db = get_db()
    rows = db.execute("SELECT code, name FROM departments ORDER BY code").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/advances", methods=["GET"])
@login_required
def list_advances():
    db = get_db()
    department = request.args.get("department")
    status = request.args.get("status")
    status_group = request.args.get("status_group")
    query = request.args.get("query", "").strip()
    sort_by = SORT_COLUMNS.get(request.args.get("sort_by", "date"), "date_given")
    sort_dir = "ASC" if request.args.get("sort_dir") == "asc" else "DESC"

    sql = "SELECT * FROM advances WHERE 1=1"
    params = []
    if department:
        sql += " AND department = ?"
        params.append(department)
    if status in ADVANCE_STATUSES:
        sql += " AND status = ?"
        params.append(status)
    elif status_group in STATUS_GROUPS:
        placeholders = ",".join("?" for _ in STATUS_GROUPS[status_group])
        sql += f" AND status IN ({placeholders})"
        params.extend(STATUS_GROUPS[status_group])
    if query:
        sql += " AND (lower(name) LIKE ? OR lower(purpose) LIKE ? OR lower(identifier) LIKE ? OR lower(voucher_no) LIKE ?)"
        like = f"%{query.lower()}%"
        params.extend([like, like, like, like])
    sql += f" ORDER BY (status='pending') DESC, {sort_by} {sort_dir}"

    rows = db.execute(sql, params).fetchall()
    return jsonify([row_to_dict(r) for r in rows])


@app.route("/api/advances", methods=["POST"])
@login_required
def create_advance():
    data = request.get_json(force=True)
    name = (data.get("name") or "").strip()
    identifier = (data.get("identifier") or "").strip()
    email = (data.get("email") or "").strip()
    department = (data.get("department") or "").strip()
    amount = data.get("amount")
    date_given = data.get("date_given")
    conf_start_date = (data.get("conf_start_date") or "").strip() or None
    conf_end_date = (data.get("conf_end_date") or "").strip() or None
    purpose = (data.get("purpose") or "").strip()

    if not name or not amount or not date_given:
        return jsonify({"error": "name, amount and date_given are required"}), 400
    if not identifier:
        return jsonify({"error": "Roll no. is required"}), 400
    if not department:
        return jsonify({"error": "Department is required"}), 400
    if not conf_start_date or not conf_end_date:
        return jsonify({"error": "Conference start date and end date are both required"}), 400
    try:
        amount = float(amount)
        if amount <= 0:
            raise ValueError
    except (TypeError, ValueError):
        return jsonify({"error": "amount must be a positive number"}), 400

    db = get_db()

    dept_row = db.execute("SELECT code FROM departments WHERE code = ?", (department,)).fetchone()
    if not dept_row:
        return jsonify({"error": f"Unknown department code '{department}'."}), 400

    if conf_end_date < conf_start_date:
        return jsonify({"error": "Conference end date can't be before the start date."}), 400

    conflict = identifier_conflict(db, identifier, name)
    if conflict:
        return (
            jsonify(
                {
                    "error": (
                        f"Roll no. '{identifier}' is already registered to {conflict['name']}. "
                        "Use that name, or correct the ID if this is a different person."
                    )
                }
            ),
            409,
        )

    blocker = student_has_pending(db, identifier)
    if blocker:
        return (
            jsonify(
                {
                    "error": "blocked",
                    "message": (
                        f"{name} already has an unsettled advance of "
                        f"₹{blocker['amount']:.2f} given on {blocker['date_given']}. "
                        "It must be settled before a new advance can be sanctioned."
                    ),
                    "existing": row_to_dict(blocker),
                }
            ),
            409,
        )

    financial_year = compute_financial_year(date_given)
    voucher_no = (data.get("voucher_no") or "").strip()
    if voucher_no:
        v_conflict = voucher_conflict(db, voucher_no)
        if v_conflict:
            return jsonify({"error": f"Voucher no. '{voucher_no}' is already used by {v_conflict['name']}."}), 409
    else:
        voucher_no = generate_voucher_no(db, financial_year)

    cur = db.execute(
        "INSERT INTO advances "
        "(identifier, name, email, department, amount, date_given, conf_start_date, conf_end_date, "
        "purpose, status, voucher_no, financial_year, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?, ?, ?)",
        (
            identifier, name, email, department or None, amount, date_given,
            conf_start_date, conf_end_date, purpose, voucher_no, financial_year,
            datetime.utcnow().isoformat(),
        ),
    )
    advance_id = cur.lastrowid
    record_history(db, advance_id, g.current_user["username"], [("created", None, voucher_no)])
    db.commit()
    row = db.execute("SELECT * FROM advances WHERE id = ?", (advance_id,)).fetchone()
    return jsonify(row_to_dict(row)), 201


@app.route("/api/advances/<int:advance_id>", methods=["PUT"])
@login_required
def update_advance(advance_id):
    data = request.get_json(force=True)
    db = get_db()
    row = db.execute("SELECT * FROM advances WHERE id = ?", (advance_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404

    fields = {}
    for f in (
        "name", "identifier", "email", "department", "amount", "date_given",
        "conf_start_date", "conf_end_date", "purpose", "voucher_no",
    ):
        if f in data:
            fields[f] = data[f]

    if not fields:
        return jsonify({"error": "no editable fields provided"}), 400

    # Status is never set through this general-purpose editor -- every
    # transition (settle / mark-paid / record-refund / reopen) goes through
    # its own dedicated endpoint below, which also captures the settlement
    # figures that transition needs.

    if "voucher_no" in fields:
        fields["voucher_no"] = (fields["voucher_no"] or "").strip()
        if not fields["voucher_no"]:
            return jsonify({"error": "Voucher no. cannot be empty"}), 400
        conflict = voucher_conflict(db, fields["voucher_no"], exclude_id=advance_id)
        if conflict:
            return jsonify({"error": f"Voucher no. '{fields['voucher_no']}' is already used by {conflict['name']}."}), 409

    if "department" in fields:
        fields["department"] = (fields["department"] or "").strip() or None
        if not fields["department"]:
            return jsonify({"error": "Department cannot be empty"}), 400
        dept_row = db.execute("SELECT code FROM departments WHERE code = ?", (fields["department"],)).fetchone()
        if not dept_row:
            return jsonify({"error": f"Unknown department code '{fields['department']}'."}), 400

    if "conf_start_date" in fields and not str(fields["conf_start_date"] or "").strip():
        return jsonify({"error": "Conference start date cannot be empty"}), 400
    if "conf_end_date" in fields and not str(fields["conf_end_date"] or "").strip():
        return jsonify({"error": "Conference end date cannot be empty"}), 400

    new_conf_start = fields.get("conf_start_date", row["conf_start_date"])
    new_conf_end = fields.get("conf_end_date", row["conf_end_date"])
    if new_conf_start and new_conf_end and new_conf_end < new_conf_start:
        return jsonify({"error": "Conference end date can't be before the start date."}), 400

    new_identifier = fields.get("identifier", row["identifier"])
    new_name = fields.get("name", row["name"])

    if "identifier" in fields and not str(new_identifier).strip():
        return jsonify({"error": "Roll no. cannot be empty"}), 400

    if "identifier" in fields or "name" in fields:
        conflict = identifier_conflict(db, new_identifier, new_name, exclude_id=advance_id)
        if conflict:
            return (
                jsonify(
                    {
                        "error": (
                            f"Roll no. '{new_identifier}' is already registered to {conflict['name']}. "
                            "Use that name, or correct the ID if this is a different person."
                        )
                    }
                ),
                409,
            )

    set_clause = ", ".join(f"{k} = ?" for k in fields)
    params = list(fields.values()) + [advance_id]
    db.execute(f"UPDATE advances SET {set_clause} WHERE id = ?", params)
    record_history(
        db, advance_id, g.current_user["username"],
        [(k, row[k], v) for k, v in fields.items()],
    )
    db.commit()
    updated = db.execute("SELECT * FROM advances WHERE id = ?", (advance_id,)).fetchone()
    return jsonify(row_to_dict(updated))


@app.route("/api/advances/<int:advance_id>/reopen", methods=["POST"])
@login_required
def reopen_advance(advance_id):
    db = get_db()
    row = db.execute("SELECT * FROM advances WHERE id = ?", (advance_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    if row["status"] == "pending":
        return jsonify({"error": "This advance is already pending."}), 400
    db.execute(
        "UPDATE advances SET status='pending', date_settled=NULL, amount_spent=NULL, "
        "amount_refundable=NULL, amount_paid=NULL, date_paid=NULL, payment_ref=NULL, "
        "amount_received=NULL, date_received=NULL, receipt_ref=NULL WHERE id = ?",
        (advance_id,),
    )
    record_history(db, advance_id, g.current_user["username"], [("status", row["status"], "pending")])
    db.commit()
    updated = db.execute("SELECT * FROM advances WHERE id = ?", (advance_id,)).fetchone()
    return jsonify(row_to_dict(updated))


def _amount_from(data, field):
    try:
        return float(data.get(field))
    except (TypeError, ValueError):
        return None


MISMATCH_TOLERANCE = 0.01

SETTLE_MESSAGES = {
    "closed": "Advance settled. Nothing further owed either way.",
    "closed_payment_due": "Advance closed. Payment due to the person.",
    "closed_refund_due": "Advance closed. Refund due from the person.",
}


@app.route("/api/advances/<int:advance_id>/settle", methods=["POST"])
@login_required
def settle_advance(advance_id):
    """pending -> closed (spent == given, nothing owed either way),
    closed_payment_due (institute owes the person) or closed_refund_due
    (person owes the institute) -- entirely driven by the amount actually
    spent vs the original advance amount. The outcome is always computed and
    returned as a preview first; nothing is written until the caller sends
    the same request back with confirm=true, even when spend matches the
    advance exactly.
    """
    db = get_db()
    row = db.execute("SELECT * FROM advances WHERE id = ?", (advance_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    if row["status"] != "pending":
        return jsonify({"error": "Only a pending advance can be settled."}), 400

    data = request.get_json(force=True)
    amount_spent = _amount_from(data, "amount_spent")
    if amount_spent is None or amount_spent < 0:
        return jsonify({"error": "amount_spent must be a non-negative number"}), 400
    confirm = bool(data.get("confirm"))

    balance = amount_spent - row["amount"]  # >0: payment due TO the person, <0: refund due FROM them
    if abs(balance) <= MISMATCH_TOLERANCE:
        new_status, amount_refundable = "closed", None
        preview = f"₹{amount_spent:.2f} spent matches the ₹{row['amount']:.2f} advance exactly -- nothing owed either way."
    elif balance > 0:
        new_status, amount_refundable = "closed_payment_due", None
        preview = f"₹{amount_spent:.2f} spent against a ₹{row['amount']:.2f} advance -- ₹{balance:.2f} will be due to {row['name']}."
    else:
        new_status, amount_refundable = "closed_refund_due", -balance
        preview = f"₹{amount_spent:.2f} spent against a ₹{row['amount']:.2f} advance -- ₹{amount_refundable:.2f} will be due from {row['name']}."

    if not confirm:
        return (
            jsonify({
                "confirm_required": True,
                "message": preview,
                "preview": {"status": new_status, "amount_spent": amount_spent, "amount_refundable": amount_refundable},
            }),
            409,
        )

    today = date.today().isoformat()
    db.execute(
        "UPDATE advances SET status=?, amount_spent=?, amount_refundable=?, date_settled=? WHERE id = ?",
        (new_status, amount_spent, amount_refundable, today, advance_id),
    )
    record_history(db, advance_id, g.current_user["username"], [
        ("status", row["status"], new_status),
        ("amount_spent", row["amount_spent"], amount_spent),
        ("amount_refundable", row["amount_refundable"], amount_refundable),
        ("date_settled", row["date_settled"], today),
    ])
    db.commit()
    updated = db.execute("SELECT * FROM advances WHERE id = ?", (advance_id,)).fetchone()
    result = row_to_dict(updated)
    result["message"] = SETTLE_MESSAGES[new_status]
    return jsonify(result)


@app.route("/api/advances/<int:advance_id>/mark-paid", methods=["POST"])
@login_required
def mark_paid(advance_id):
    """closed_payment_due -> payment_made: institute has paid the balance owed."""
    db = get_db()
    row = db.execute("SELECT * FROM advances WHERE id = ?", (advance_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    if row["status"] != "closed_payment_due":
        return jsonify({"error": "Only an advance with payment due can be marked paid."}), 400

    data = request.get_json(force=True)
    amount_paid = _amount_from(data, "amount_paid")
    date_paid = (data.get("date_paid") or "").strip()
    payment_ref = (data.get("payment_ref") or "").strip()
    confirm = bool(data.get("confirm"))
    if amount_paid is None or amount_paid <= 0:
        return jsonify({"error": "amount_paid must be a positive number"}), 400
    if not date_paid:
        return jsonify({"error": "date_paid is required"}), 400

    expected = (row["amount_spent"] or 0) - row["amount"]
    if not confirm and abs(amount_paid - expected) > MISMATCH_TOLERANCE:
        return (
            jsonify({
                "mismatch": True,
                "message": f"You entered ₹{amount_paid:.2f}, but the expected payment is ₹{expected:.2f}.",
            }),
            409,
        )

    db.execute(
        "UPDATE advances SET status='payment_made', amount_paid=?, date_paid=?, payment_ref=? WHERE id = ?",
        (amount_paid, date_paid, payment_ref or None, advance_id),
    )
    record_history(db, advance_id, g.current_user["username"], [
        ("status", row["status"], "payment_made"),
        ("amount_paid", row["amount_paid"], amount_paid),
        ("date_paid", row["date_paid"], date_paid),
        ("payment_ref", row["payment_ref"], payment_ref or None),
    ])
    db.commit()
    updated = db.execute("SELECT * FROM advances WHERE id = ?", (advance_id,)).fetchone()
    return jsonify(row_to_dict(updated))


@app.route("/api/advances/<int:advance_id>/record-refund", methods=["POST"])
@login_required
def record_refund(advance_id):
    """closed_refund_due -> refund_received: person has paid back the shortfall."""
    db = get_db()
    row = db.execute("SELECT * FROM advances WHERE id = ?", (advance_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    if row["status"] != "closed_refund_due":
        return jsonify({"error": "Only an advance with a refund due can have a refund recorded."}), 400

    data = request.get_json(force=True)
    amount_received = _amount_from(data, "amount_received")
    date_received = (data.get("date_received") or "").strip()
    receipt_ref = (data.get("receipt_ref") or "").strip()
    confirm = bool(data.get("confirm"))
    if amount_received is None or amount_received <= 0:
        return jsonify({"error": "amount_received must be a positive number"}), 400
    if not date_received:
        return jsonify({"error": "date_received is required"}), 400

    expected = row["amount_refundable"] or 0
    if not confirm and abs(amount_received - expected) > MISMATCH_TOLERANCE:
        return (
            jsonify({
                "mismatch": True,
                "message": f"You entered ₹{amount_received:.2f}, but the refund due is ₹{expected:.2f}.",
            }),
            409,
        )

    db.execute(
        "UPDATE advances SET status='refund_received', amount_received=?, date_received=?, receipt_ref=? WHERE id = ?",
        (amount_received, date_received, receipt_ref or None, advance_id),
    )
    record_history(db, advance_id, g.current_user["username"], [
        ("status", row["status"], "refund_received"),
        ("amount_received", row["amount_received"], amount_received),
        ("date_received", row["date_received"], date_received),
        ("receipt_ref", row["receipt_ref"], receipt_ref or None),
    ])
    db.commit()
    updated = db.execute("SELECT * FROM advances WHERE id = ?", (advance_id,)).fetchone()
    return jsonify(row_to_dict(updated))


@app.route("/api/advances/<int:advance_id>/history", methods=["GET"])
@login_required
def advance_history(advance_id):
    db = get_db()
    row = db.execute("SELECT id FROM advances WHERE id = ?", (advance_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    rows = db.execute(
        "SELECT * FROM advance_history WHERE advance_id = ? ORDER BY id ASC", (advance_id,)
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/activity", methods=["GET"])
@login_required
def recent_activity():
    """
    Dashboard's recent-activity feed: the last N history events across every
    advance in the register (not just one), newest first, with enough
    context (name, voucher no) to read standalone without following a link.
    A left join so a since-deleted advance's history still shows up rather
    than silently vanishing.
    """
    limit = request.args.get("limit", 20, type=int)
    limit = max(1, min(limit, 100))
    db = get_db()
    rows = db.execute(
        """
        SELECT h.id, h.advance_id, h.field, h.old_value, h.new_value, h.changed_by, h.changed_at,
               a.name AS advance_name, a.voucher_no AS advance_voucher_no
        FROM advance_history h
        LEFT JOIN advances a ON a.id = h.advance_id
        ORDER BY h.id DESC
        LIMIT ?
        """,
        (limit,),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/advances/<int:advance_id>", methods=["DELETE"])
@login_required
def delete_advance(advance_id):
    db = get_db()
    db.execute("DELETE FROM advances WHERE id = ?", (advance_id,))
    db.execute("DELETE FROM advance_history WHERE advance_id = ?", (advance_id,))
    db.commit()
    return jsonify({"deleted": True})


# ---------------------------------------------------------------------------
# Email reminders — preview first, send only after explicit confirmation
# ---------------------------------------------------------------------------

def build_reminder(row):
    d = row_to_dict(row)
    days = d.get("days_outstanding", 0)
    overdue_line = ""
    if d.get("is_overdue"):
        overdue_line = (
            f" This is now {d['days_overdue']} day{'s' if d['days_overdue'] != 1 else ''} past the "
            f"settlement deadline of {SETTLEMENT_GRACE_DAYS} days after the conference end date "
            f"({d.get('due_date')})."
        )
    subject = f"Reminder: Advance of ₹{row['amount']:.2f} pending for {days} day{'s' if days != 1 else ''}"
    body = (
        f"Dear {row['name']},\n\n"
        f"This is a reminder that an advance of ₹{row['amount']:.2f} was issued to you on "
        f"{row['date_given']}"
        + (f" for {row['purpose']}" if row["purpose"] else "")
        + f" and remains unsettled, {days} day{'s' if days != 1 else ''} after it was given."
        + overdue_line
        + "\n\nKindly submit your settlement details or contact the accounts office at the earliest.\n\n"
        f"Regards,\nAccounts Office"
    )
    return subject, body


@app.route("/api/advances/<int:advance_id>/remind/preview", methods=["GET"])
@login_required
def preview_reminder(advance_id):
    db = get_db()
    row = db.execute("SELECT * FROM advances WHERE id = ?", (advance_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    if row["status"] != "pending":
        return jsonify({"error": "Reminders are only for advances that are still pending settlement."}), 400
    if not row["email"]:
        return jsonify({"error": "This person has no email address on file."}), 400
    subject, body = build_reminder(row)
    return jsonify({"to": row["email"], "subject": subject, "body": body})


# ---------------------------------------------------------------------------
# Shared SMTP helper
# ---------------------------------------------------------------------------

def smtp_configured():
    return all([os.environ.get("SMTP_HOST"), os.environ.get("SMTP_USER"), os.environ.get("SMTP_PASS")])


def send_email(to_email, subject, body):
    """Returns (ok: bool, error_message: str|None)."""
    smtp_host = os.environ.get("SMTP_HOST")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    smtp_user = os.environ.get("SMTP_USER")
    smtp_pass = os.environ.get("SMTP_PASS")
    from_email = os.environ.get("FROM_EMAIL", smtp_user)

    if not smtp_configured():
        return False, (
            "Email is not configured yet. Set SMTP_HOST, SMTP_USER and SMTP_PASS "
            "in backend/.env (see README) and restart the server."
        )

    msg = MIMEText(body)
    msg["Subject"] = subject
    msg["From"] = from_email
    msg["To"] = to_email

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(from_email, [to_email], msg.as_string())
    except Exception as e:
        return False, f"Failed to send email: {e}"
    return True, None


@app.route("/api/advances/<int:advance_id>/remind/send", methods=["POST"])
@login_required
def send_reminder(advance_id):
    db = get_db()
    row = db.execute("SELECT * FROM advances WHERE id = ?", (advance_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    if row["status"] != "pending":
        return jsonify({"error": "Reminders are only for advances that are still pending settlement."}), 400
    if not row["email"]:
        return jsonify({"error": "This person has no email address on file."}), 400

    data = request.get_json(force=True)
    subject = data.get("subject", "").strip()
    body = data.get("body", "").strip()
    if not subject or not body:
        return jsonify({"error": "subject and body are required"}), 400

    ok, err = send_email(row["email"], subject, body)
    if not ok:
        return jsonify({"error": err}), 500 if "not configured" in err else 502

    db.execute(
        "UPDATE advances SET last_reminded_at = ? WHERE id = ?",
        (datetime.utcnow().isoformat(), advance_id),
    )
    db.commit()
    return jsonify({"sent": True, "to": row["email"], "subject": subject})


@app.route("/api/advances/remind/bulk", methods=["POST"])
@login_required
def send_reminders_bulk():
    """
    Sends the standard reminder to every pending advance that's overdue,
    has an email on file, and hasn't been reminded (by this or the automated
    job) within REMINDER_COOLDOWN_DAYS. Used by the "Remind all overdue"
    button on the dashboard, and shares its cooldown logic with the
    standalone scheduled script (send_reminders.py) so a person isn't
    double-emailed just because both paths ran the same day.
    """
    cooldown_days = int(os.environ.get("REMINDER_COOLDOWN_DAYS", "7"))
    db = get_db()
    cutoff = (datetime.utcnow() - timedelta(days=cooldown_days)).isoformat()
    rows = db.execute(
        """
        SELECT * FROM advances
        WHERE status = 'pending'
          AND email IS NOT NULL AND trim(email) != ''
          AND conf_end_date IS NOT NULL
          AND date(conf_end_date, ?) < date('now')
          AND (last_reminded_at IS NULL OR last_reminded_at < ?)
        """,
        (f"+{SETTLEMENT_GRACE_DAYS} days", cutoff),
    ).fetchall()

    sent, failed = [], []
    for row in rows:
        subject, body = build_reminder(row)
        ok, err = send_email(row["email"], subject, body)
        if ok:
            db.execute(
                "UPDATE advances SET last_reminded_at = ? WHERE id = ?",
                (datetime.utcnow().isoformat(), row["id"]),
            )
            sent.append({"id": row["id"], "name": row["name"], "to": row["email"]})
        else:
            failed.append({"id": row["id"], "name": row["name"], "error": err})
    db.commit()
    return jsonify({"sent": sent, "failed": failed, "checked": len(rows)})


# ---------------------------------------------------------------------------
# Excel import
# ---------------------------------------------------------------------------

HEADER_ALIASES = {
    "identifier": "identifier", "roll_no": "identifier", "rollno": "identifier",
    "roll no": "identifier", "roll number": "identifier", "id": "identifier",
    "name": "name",
    "email": "email", "email address": "email",
    "department": "department", "dept": "department", "department code": "department",
    "amount": "amount",
    "date_given": "date_given", "date given": "date_given", "date": "date_given",
    "conf_start_date": "conf_start_date", "conference start": "conf_start_date",
    "conf start date": "conf_start_date", "conference start date": "conf_start_date",
    "conf_end_date": "conf_end_date", "conference end": "conf_end_date",
    "conf end date": "conf_end_date", "conference end date": "conf_end_date",
    "purpose": "purpose",
    "voucher_no": "voucher_no", "voucher no": "voucher_no", "voucher number": "voucher_no",
    "voucher": "voucher_no",
}


def normalize_header(h):
    if h is None:
        return None
    key = str(h).strip().lower()
    key = re.sub(r"[.:]+$", "", key)          # drop trailing punctuation like "Roll No."
    key = re.sub(r"\s+", " ", key).strip()     # collapse repeated spaces
    key_underscored = key.replace(" ", "_").replace("-", "_")
    return (
        HEADER_ALIASES.get(key)
        or HEADER_ALIASES.get(key_underscored)
        or HEADER_ALIASES.get(key.replace("-", " "))
    )


def normalize_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d") if isinstance(value, datetime) else value.isoformat()
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


@app.route("/api/import", methods=["POST"])
@login_required
def import_excel():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
    except Exception as e:
        return jsonify({"error": f"Could not read this file as an Excel workbook: {e}"}), 400

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return jsonify({"error": "The sheet needs a header row plus at least one data row."}), 400

    headers = [normalize_header(h) for h in rows[0]]
    if "identifier" not in headers or "name" not in headers or "amount" not in headers or "date_given" not in headers:
        return (
            jsonify(
                {
                    "error": (
                        "The sheet must have columns for roll no, name, amount and date given. "
                        "Voucher No. is optional -- leave it blank to auto-generate. "
                        "See the README for the exact column names accepted."
                    )
                }
            ),
            400,
        )

    db = get_db()
    inserted, updated, skipped = 0, 0, []

    for i, raw_row in enumerate(rows[1:], start=2):
        record = dict(zip(headers, raw_row))
        record.pop(None, None)

        identifier = str(record.get("identifier") or "").strip()
        name = str(record.get("name") or "").strip()
        amount = record.get("amount")
        date_given = normalize_date(record.get("date_given"))
        email = str(record.get("email") or "").strip()
        department = str(record.get("department") or "").strip().upper() or None
        conf_start_date = normalize_date(record.get("conf_start_date"))
        conf_end_date = normalize_date(record.get("conf_end_date"))
        purpose = str(record.get("purpose") or "").strip()
        voucher_no_in = str(record.get("voucher_no") or "").strip() or None

        if not identifier or not name or amount in (None, "") or not date_given:
            skipped.append({"row": i, "reason": "missing roll no, name, amount, or a valid date given"})
            continue
        try:
            amount = float(amount)
            if amount <= 0:
                raise ValueError
        except (TypeError, ValueError):
            skipped.append({"row": i, "reason": "amount is not a valid positive number"})
            continue

        if department:
            dept_row = db.execute("SELECT code FROM departments WHERE code = ?", (department,)).fetchone()
            if not dept_row:
                skipped.append({"row": i, "reason": f"unknown department code '{department}'"})
                continue

        conflict = identifier_conflict(db, identifier, name)
        if conflict:
            skipped.append(
                {
                    "row": i,
                    "reason": f"'{identifier}' is already registered to {conflict['name']}, not {name}",
                }
            )
            continue

        existing = db.execute(
            "SELECT * FROM advances WHERE lower(trim(identifier)) = lower(trim(?)) "
            "AND lower(trim(name)) = lower(trim(?)) AND status = 'pending'",
            (identifier, name),
        ).fetchone()

        if voucher_no_in:
            v_conflict = voucher_conflict(db, voucher_no_in, exclude_id=existing["id"] if existing else None)
            if v_conflict:
                skipped.append({"row": i, "reason": f"voucher no. '{voucher_no_in}' is already used by {v_conflict['name']}"})
                continue

        if existing:
            if voucher_no_in:
                db.execute(
                    "UPDATE advances SET amount=?, date_given=?, email=?, department=?, conf_start_date=?, "
                    "conf_end_date=?, purpose=?, voucher_no=? WHERE id=?",
                    (amount, date_given, email, department, conf_start_date, conf_end_date,
                     purpose, voucher_no_in, existing["id"]),
                )
            else:
                db.execute(
                    "UPDATE advances SET amount=?, date_given=?, email=?, department=?, conf_start_date=?, "
                    "conf_end_date=?, purpose=? WHERE id=?",
                    (amount, date_given, email, department, conf_start_date, conf_end_date,
                     purpose, existing["id"]),
                )
            updated += 1
        else:
            blocker = student_has_pending(db, identifier)
            if blocker:
                skipped.append(
                    {
                        "row": i,
                        "reason": f"{name} already has a pending advance on file; settle it first",
                    }
                )
                continue
            financial_year = compute_financial_year(date_given)
            voucher_no = voucher_no_in or generate_voucher_no(db, financial_year)
            db.execute(
                "INSERT INTO advances (identifier, name, email, department, amount, date_given, "
                "conf_start_date, conf_end_date, purpose, status, voucher_no, financial_year, created_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?, ?, ?)",
                (identifier, name, email, department, amount, date_given, conf_start_date,
                 conf_end_date, purpose, voucher_no, financial_year,
                 datetime.utcnow().isoformat()),
            )
            inserted += 1

    db.commit()
    return jsonify({"inserted": inserted, "updated": updated, "skipped": skipped})


# ---------------------------------------------------------------------------
# Reports (Excel / PDF export)
# ---------------------------------------------------------------------------

REPORT_STATUS_LABELS = {
    "pending": "Pending",
    "closed": "Settled (nothing owed)",
    "closed_payment_due": "Payment due",
    "payment_made": "Paid",
    "closed_refund_due": "Refund due",
    "refund_received": "Refunded",
}


def _report_rows(db):
    """Shared filter logic for both report exports: department, status or
    status_group, a date_given range, and free-text search -- the same
    vocabulary as the Advances tab, plus date_from/date_to."""
    department = request.args.get("department")
    status = request.args.get("status")
    status_group = request.args.get("status_group")
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    query = request.args.get("query", "").strip()

    sql = "SELECT * FROM advances WHERE 1=1"
    params = []
    if department:
        sql += " AND department = ?"
        params.append(department)
    if status in ADVANCE_STATUSES:
        sql += " AND status = ?"
        params.append(status)
    elif status_group in STATUS_GROUPS:
        placeholders = ",".join("?" for _ in STATUS_GROUPS[status_group])
        sql += f" AND status IN ({placeholders})"
        params.extend(STATUS_GROUPS[status_group])
    if date_from:
        sql += " AND date_given >= ?"
        params.append(date_from)
    if date_to:
        sql += " AND date_given <= ?"
        params.append(date_to)
    if query:
        sql += " AND (lower(name) LIKE ? OR lower(purpose) LIKE ? OR lower(identifier) LIKE ? OR lower(voucher_no) LIKE ?)"
        like = f"%{query.lower()}%"
        params.extend([like, like, like, like])
    sql += " ORDER BY date_given DESC, id DESC"
    return [row_to_dict(r) for r in db.execute(sql, params).fetchall()]


def _report_filename(prefix, ext):
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return f"{prefix}_{stamp}.{ext}"


def _owed_amount(a):
    """What's still outstanding on this row, whichever direction it runs."""
    if a["status"] == "pending":
        return a["amount"]
    if a["status"] == "closed_payment_due":
        return (a["amount_spent"] or 0) - a["amount"]
    if a["status"] == "closed_refund_due":
        return a["amount_refundable"] or 0
    return 0


@app.route("/api/reports/excel", methods=["GET"])
@login_required
def export_excel_report():
    db = get_db()
    rows = _report_rows(db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Advance Register"

    headers = [
        "Voucher No.", "Roll No.", "Name", "Department", "Amount", "Date Given",
        "Conf. Start", "Conf. End", "Status", "Amount Spent", "Amount Refundable",
        "Amount Paid", "Amount Received", "Date Settled", "Days Outstanding / Overdue",
    ]
    header_fill = PatternFill(start_color="6D28D9", end_color="6D28D9", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for r, a in enumerate(rows, start=2):
        overdue_note = f"{a['days_overdue']}d overdue" if a.get("is_overdue") else (
            f"{a.get('days_outstanding', '')}d" if a.get("days_outstanding") is not None else ""
        )
        ws.append([
            a["voucher_no"], a["identifier"], a["name"], a["department"], a["amount"],
            a["date_given"], a["conf_start_date"], a["conf_end_date"],
            REPORT_STATUS_LABELS.get(a["status"], a["status"]),
            a["amount_spent"], a["amount_refundable"], a["amount_paid"], a["amount_received"],
            a["date_settled"], overdue_note,
        ])

    widths = [16, 14, 24, 12, 12, 12, 12, 12, 18, 13, 15, 12, 15, 13, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Advance Register -- Summary", ""])
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.append(["Generated", datetime.now().strftime("%d-%b-%Y %H:%M")])
    ws2.append(["Rows in this report", len(rows)])
    ws2.append([])
    ws2.append(["Status", "Count", "Total amount"])
    for c in ws2["A5:C5"][0]:
        c.font = Font(bold=True)
    totals = {}
    for a in rows:
        totals.setdefault(a["status"], {"count": 0, "amount": 0.0})
        totals[a["status"]]["count"] += 1
        totals[a["status"]]["amount"] += a["amount"] or 0
    for status, t in totals.items():
        ws2.append([REPORT_STATUS_LABELS.get(status, status), t["count"], round(t["amount"], 2)])
    for col, w in zip("ABC", (28, 10, 16)):
        ws2.column_dimensions[col].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name=_report_filename("advance_register", "xlsx"),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/reports/pdf", methods=["GET"])
@login_required
def export_pdf_report():
    db = get_db()
    rows = _report_rows(db)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=16 * mm, rightMargin=16 * mm, topMargin=14 * mm, bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleV", parent=styles["Title"], textColor=colors.HexColor("#4C1D95"), fontSize=18)
    sub_style = ParagraphStyle("SubV", parent=styles["Normal"], textColor=colors.HexColor("#635C7C"), fontSize=9)

    filters_desc = []
    if request.args.get("department"):
        filters_desc.append(f"Department: {request.args.get('department')}")
    if request.args.get("status"):
        filters_desc.append(f"Status: {REPORT_STATUS_LABELS.get(request.args.get('status'), request.args.get('status'))}")
    if request.args.get("status_group"):
        filters_desc.append(f"View: {request.args.get('status_group').replace('_', ' ').title()}")
    if request.args.get("date_from") or request.args.get("date_to"):
        filters_desc.append(f"Date given: {request.args.get('date_from') or '…'} to {request.args.get('date_to') or '…'}")
    if request.args.get("query"):
        filters_desc.append(f"Search: \"{request.args.get('query')}\"")

    story = [
        Paragraph("Advance Register", title_style),
        Paragraph(
            f"Generated {datetime.now().strftime('%d-%b-%Y %H:%M')} &nbsp;·&nbsp; {len(rows)} record(s)"
            + (" &nbsp;·&nbsp; " + " | ".join(filters_desc) if filters_desc else ""),
            sub_style,
        ),
        Spacer(1, 10),
    ]

    table_header = ["Voucher No.", "Roll No.", "Name", "Dept.", "Amount", "Date Given", "Status", "Outstanding / Due"]
    table_data = [table_header]
    for a in rows:
        table_data.append([
            a["voucher_no"] or "", a["identifier"] or "", a["name"] or "", a["department"] or "",
            f"Rs {a['amount']:,.2f}", a["date_given"] or "",
            REPORT_STATUS_LABELS.get(a["status"], a["status"]),
            f"Rs {_owed_amount(a):,.2f}" if _owed_amount(a) else "--",
        ])

    if len(table_data) == 1:
        story.append(Paragraph("No advances match these filters.", styles["Normal"]))
    else:
        col_widths = [28 * mm, 22 * mm, 45 * mm, 18 * mm, 26 * mm, 24 * mm, 32 * mm, 30 * mm]
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6D28D9")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F4FE")]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E8E3F7")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (4, 1), (4, -1), "RIGHT"),
            ("ALIGN", (7, 1), (7, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(t)

    doc.build(story)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name=_report_filename("advance_register", "pdf"),
        mimetype="application/pdf",
    )


@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


# init_db() runs at import time, not just under __main__, so a production
# WSGI server (gunicorn) importing this module also gets the schema created/
# migrated -- gunicorn never executes the __main__ block below.
init_db()


if __name__ == "__main__":
    # debug=True (the interactive debugger + auto-reload) must never run
    # outside local development -- it can execute arbitrary code if reached
    # over the network. FLASK_DEBUG must be explicitly set to enable it.
    debug_mode = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="127.0.0.1", port=int(os.environ.get("PORT", 5000)), debug=debug_mode)
